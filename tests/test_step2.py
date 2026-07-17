"""
Tests for Step 2: Billing module for Sample data
"""

import os
import sys
import tempfile
from datetime import datetime, date
import pytest

# Add project directory to path
sys.path.insert(0, os.path.dirname(os.path.abspath(__file__)) + '/..')

from flask import Flask
from app.extensions import db
from app.models import Sample, FSO
from app.billing.billing_utils import compute_summary, format_price, generate_excel_report


@pytest.fixture
def app():
    """Create and configure a test Flask app."""
    app = Flask(__name__)
    
    # Use in-memory database for tests
    app.config['SQLALCHEMY_DATABASE_URI'] = 'sqlite:///:memory:'
    app.config['SQLALCHEMY_TRACK_MODIFICATIONS'] = False
    app.config['TESTING'] = True
    
    # Register billing blueprint
    from app.billing.routes import billing_bp
    app.register_blueprint(billing_bp, url_prefix='/billing')
    
    with app.app_context():
        db.init_app(app)
        db.create_all()
        yield app
        db.drop_all()


class TestBillingUtils:
    """Tests for billing utility functions."""
    
    def test_format_price_basic(self, app):
        """Test basic price formatting."""
        with app.app_context():
            assert format_price("100") == 100.0
            assert format_price("100.50") == 100.50
            assert format_price("₹100") == 100.0
            assert format_price("Rs 100") == 100.0
            assert format_price("1,000") == 1000.0
            assert format_price("1,000.50") == 1000.50
    
    def test_format_price_none(self, app):
        """Test formatting of None/empty values."""
        with app.app_context():
            assert format_price(None) == 0.0
            assert format_price("") == 0.0
            assert format_price("  ") == 0.0
    
    def test_format_price_invalid(self, app):
        """Test formatting of invalid values."""
        with app.app_context():
            assert format_price("invalid") == 0.0
            assert format_price("abc123") == 0.0
    
    def test_compute_summary_basic(self, app):
        """Test summary computation with basic data."""
        with app.app_context():
            # Create mock sample data (as dicts for simplicity)
            samples = [
                {'sample_type': 'Food', 'price': '100.00'},
                {'sample_type': 'Food', 'price': '150.00'},
                {'sample_type': 'Water', 'price': '50.00'},
            ]
            
            summary = compute_summary(samples)
            
            assert summary['total_count'] == 3
            assert summary['grand_total'] == 300.0
            assert 'Food' in summary['by_type']
            assert 'Water' in summary['by_type']
            assert summary['by_type']['Food']['count'] == 2
            assert summary['by_type']['Food']['total_price'] == 250.0
            assert summary['by_type']['Water']['count'] == 1
            assert summary['by_type']['Water']['total_price'] == 50.0
    
    def test_compute_summary_with_models(self, app):
        """Test summary computation with Sample model objects."""
        with app.app_context():
            # First create an FSO
            fso = FSO(fso_name="Test FSO")
            db.session.add(fso)
            
            # Create sample objects
            samples = [
                Sample(
                    sample_code="SKS-2026-00001",
                    sample_name="Sample 1",
                    sample_type="Food",
                    fso_name="Test FSO",
                    collection_date="2026-07-17",
                    price="100.00"
                ),
                Sample(
                    sample_code="SKS-2026-00002",
                    sample_name="Sample 2",
                    sample_type="Food",
                    fso_name="Test FSO",
                    collection_date="2026-07-17",
                    price="150.50"
                ),
                Sample(
                    sample_code="SKS-2026-00003",
                    sample_name="Sample 3",
                    sample_type="Water",
                    fso_name="Test FSO",
                    collection_date="2026-07-17",
                    price="75.25"
                )
            ]
            
            for sample in samples:
                db.session.add(sample)
            db.session.commit()
            
            # Query and compute summary
            all_samples = Sample.query.all()
            summary = compute_summary(all_samples)
            
            assert summary['total_count'] == 3
            assert abs(summary['grand_total'] - 325.75) < 0.01
            assert summary['by_type']['Food']['count'] == 2
            assert abs(summary['by_type']['Food']['total_price'] - 250.50) < 0.01
            assert summary['by_type']['Water']['count'] == 1
            assert abs(summary['by_type']['Water']['total_price'] - 75.25) < 0.01
    
    def test_compute_summary_empty(self, app):
        """Test summary computation with empty list."""
        with app.app_context():
            summary = compute_summary([])
            assert summary['total_count'] == 0
            assert summary['grand_total'] == 0.0
            assert summary['by_type'] == {}
    
    def test_compute_summary_null_types(self, app):
        """Test summary computation handles null/empty sample_type."""
        with app.app_context():
            samples = [
                {'sample_type': None, 'price': '100'},
                {'sample_type': '', 'price': '50'},
                {'sample_type': 'Food', 'price': '75'},
            ]
            
            summary = compute_summary(samples)
            
            # None and empty string should be grouped as 'Other'
            assert summary['total_count'] == 3
            assert summary['grand_total'] == 225.0
            assert summary['by_type']['Other']['count'] == 2
            assert summary['by_type']['Other']['total_price'] == 150.0
            assert summary['by_type']['Food']['count'] == 1
            assert summary['by_type']['Food']['total_price'] == 75.0


class TestExcelExport:
    """Tests for Excel export functionality."""
    
    def test_excel_export_basic(self, app):
        """Test Excel export generates valid file."""
        with app.app_context():
            samples = [
                {'sample_code': 'SKS-2026-00001', 'sample_name': 'Test', 
                 'sample_type': 'Food', 'fso_name': 'FSO1', 
                 'collection_date': '2026-07-17', 'price': '100.00'}
            ]
            
            summary = compute_summary(samples)
            
            excel_file, filename = generate_excel_report(samples, summary)
            
            # Verify file was generated
            assert excel_file is not None
            assert filename == "billing_summary_all.xlsx"
            
            # Verify it's valid Excel by checking file size
            content = excel_file.getvalue()
            assert len(content) > 1000  # Excel file should be at least 1KB
    
    def test_excel_export_with_dates(self, app):
        """Test Excel export with date range in filename."""
        with app.app_context():
            samples = []
            summary = {'by_type': {}, 'grand_total': 0, 'total_count': 0}
            
            excel_file, filename = generate_excel_report(
                samples, summary, 
                start_date='2026-07-01', 
                end_date='2026-07-31'
            )
            
            assert '2026-07-01' in filename
            assert '2026-07-31' in filename
            assert 'to' in filename
    
    def test_excel_export_structure(self, app):
        """Test Excel file has correct sheet structure."""
        import openpyxl
        
        with app.app_context():
            samples = [
                {'sample_code': 'SKS-2026-00001', 'sample_name': 'Test1',
                 'sample_type': 'Food', 'fso_name': 'FSO1',
                 'collection_date': '2026-07-17', 'submission_date': '2026-07-18',
                 'retailer_fssai': '12345', 'retailer_name': 'Retailer1', 'price': '100.00'},
                {'sample_code': 'SKS-2026-00002', 'sample_name': 'Test2',
                 'sample_type': 'Water', 'fso_name': 'FSO2',
                 'collection_date': '2026-07-17', 'submission_date': None,
                 'retailer_fssai': '67890', 'retailer_name': None, 'price': '50.00'}
            ]
            
            summary = compute_summary(samples)
            
            excel_file, filename = generate_excel_report(samples, summary)
            
            # Load the Excel file and verify structure
            wb = openpyxl.load_workbook(excel_file)
            
            # Check sheets exist
            assert 'Samples' in wb.sheetnames
            assert 'Summary' in wb.sheetnames
            
            # Check Samples sheet
            ws_samples = wb['Samples']
            # Should have header row + 2 data rows
            assert ws_samples.max_row >= 3
            
            # Check header
            headers = [cell.value for cell in ws_samples[1]]
            assert 'Sample Code' in headers
            assert 'Sample Name' in headers
            assert 'Sample Type' in headers
            
            # Check Summary sheet
            ws_summary = wb['Summary']
            # Should have header, filter info, summary table, grand total
            assert ws_summary.max_row >= 6
            
            # Check for billing summary text
            assert 'Billing Summary' in str(ws_summary.cell(row=2, column=1).value)


class TestBillingRoutes:
    """Tests for Billing routes."""
    
    def test_billing_blueprint_exists(self):
        """Test that billing blueprint can be imported."""
        from app.billing.routes import billing_bp
        assert billing_bp is not None
        assert billing_bp.name == 'billing'
    
    def test_billing_routes_module_import(self):
        """Test that billing routes module can be imported."""
        from app.billing import routes
        assert routes.billing_bp is not None
    
    def test_billing_utils_import(self):
        """Test that billing utils can be imported."""
        from app.billing import billing_utils
        assert hasattr(billing_utils, 'compute_summary')
        assert hasattr(billing_utils, 'generate_excel_report')
        assert hasattr(billing_utils, 'format_price')


class TestBillingFilters:
    """Tests for billing filter functionality."""
    
    def test_date_filter(self, app):
        """Test filtering by date range."""
        from app.models import Sample, FSO
        from app.billing.billing_utils import compute_summary
        
        with app.app_context():
            # Add test data
            fso = FSO(fso_name="Test FSO")
            db.session.add(fso)
            
            sample1 = Sample(
                sample_code="SKS-2026-00001",
                sample_name="Sample 1",
                sample_type="Food",
                fso_name="Test FSO",
                collection_date="2026-07-17",
                price="100.00"
            )
            sample2 = Sample(
                sample_code="SKS-2026-00002",
                sample_name="Sample 2",
                sample_type="Water",
                fso_name="Test FSO",
                collection_date="2026-07-18",
                price="50.00"
            )
            db.session.add(sample1)
            db.session.add(sample2)
            db.session.commit()
            
            # All samples
            all_samples = Sample.query.filter(
                Sample.collection_date >= '2026-07-17',
                Sample.collection_date <= '2026-07-18'
            ).all()
            summary = compute_summary(all_samples)
            assert summary['total_count'] == 2
            
            # Only 2026-07-17
            filtered_samples = Sample.query.filter(
                Sample.collection_date == '2026-07-17'
            ).all()
            summary = compute_summary(filtered_samples)
            assert summary['total_count'] == 1
    
    def test_fso_filter(self, app):
        """Test filtering by FSO name."""
        from app.models import Sample, FSO
        from app.billing.billing_utils import compute_summary
        
        with app.app_context():
            # Add FSOs
            fso1 = FSO(fso_name="FSO 1")
            fso2 = FSO(fso_name="FSO 2")
            db.session.add(fso1)
            db.session.add(fso2)
            
            # Add samples
            sample1 = Sample(
                sample_code="SKS-2026-00001", sample_name="S1",
                sample_type="Food", fso_name="FSO 1",
                collection_date="2026-07-17", price="100"
            )
            sample2 = Sample(
                sample_code="SKS-2026-00002", sample_name="S2",
                sample_type="Water", fso_name="FSO 2",
                collection_date="2026-07-17", price="50"
            )
            db.session.add(sample1)
            db.session.add(sample2)
            db.session.commit()
            
            # Filter by FSO 1
            filtered = Sample.query.filter(Sample.fso_name == 'FSO 1').all()
            summary = compute_summary(filtered)
            assert summary['total_count'] == 1
            assert summary['by_type']['Food']['count'] == 1
    
    def test_sample_type_filter(self, app):
        """Test filtering by sample type."""
        from app.models import Sample, FSO
        from app.billing.billing_utils import compute_summary
        
        with app.app_context():
            fso = FSO(fso_name="Test FSO")
            db.session.add(fso)
            
            samples = [
                Sample(sample_code="SKS-2026-00001", sample_name="S1",
                       sample_type="Food", fso_name="Test FSO",
                       collection_date="2026-07-17", price="100"),
                Sample(sample_code="SKS-2026-00002", sample_name="S2",
                       sample_type="Water", fso_name="Test FSO",
                       collection_date="2026-07-17", price="50"),
                Sample(sample_code="SKS-2026-00003", sample_name="S3",
                       sample_type="Food", fso_name="Test FSO",
                       collection_date="2026-07-17", price="75"),
            ]
            for s in samples:
                db.session.add(s)
            db.session.commit()
            
            # Filter by Food type
            filtered = Sample.query.filter(Sample.sample_type == 'Food').all()
            summary = compute_summary(filtered)
            assert summary['total_count'] == 2
            assert summary['by_type']['Food']['count'] == 2


class TestGroupingCorrectness:
    """Tests for grouping/sum correctness as specified in requirements."""
    
    def test_grouping_and_sum_correctness(self, app):
        """
        Test that grouping and sum calculations are correct.
        
        Fixture dataset:
        - 3 sample types: Food, Water, Oil
        - Verify totals match manual calculation
        """
        with app.app_context():
            # Create fixture data
            samples = [
                # Food samples
                {'sample_type': 'Food', 'price': '100.00'},
                {'sample_type': 'Food', 'price': '150.50'},
                {'sample_type': 'Food', 'price': '200.25'},
                
                # Water samples
                {'sample_type': 'Water', 'price': '50.00'},
                {'sample_type': 'Water', 'price': '75.75'},
                
                # Oil samples
                {'sample_type': 'Oil', 'price': '300.00'},
            ]
            
            summary = compute_summary(samples)
            
            # Manual calculations:
            # Food: 100 + 150.50 + 200.25 = 450.75, count = 3
            # Water: 50 + 75.75 = 125.75, count = 2
            # Oil: 300, count = 1
            # Grand total: 450.75 + 125.75 + 300 = 876.50
            # Total count: 6
            
            assert summary['total_count'] == 6
            assert abs(summary['grand_total'] - 876.50) < 0.01
            
            assert summary['by_type']['Food']['count'] == 3
            assert abs(summary['by_type']['Food']['total_price'] - 450.75) < 0.01
            
            assert summary['by_type']['Water']['count'] == 2
            assert abs(summary['by_type']['Water']['total_price'] - 125.75) < 0.01
            
            assert summary['by_type']['Oil']['count'] == 1
            assert abs(summary['by_type']['Oil']['total_price'] - 300.00) < 0.01


if __name__ == '__main__':
    import pytest
    pytest.main([__file__, '-v'])
