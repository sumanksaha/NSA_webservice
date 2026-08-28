"""Tests for the billing Excel styling extraction (SheetStyler).

Pure openpyxl — no Flask app context needed, matching the style of the
existing ``TestExcelExport`` suite in ``test_step2.py`` but targeting the
styler helpers directly so they're testable in isolation.
"""

from datetime import datetime

from openpyxl import Workbook

from app.billing.excel_styler import SheetStyler


class TestStyleHeader:
    """Header cells: bold + centered + thin border + gray fill."""

    def test_header_has_bold_font(self):
        wb = Workbook()
        cell = wb.active.cell(row=1, column=1, value="Header")
        SheetStyler().style_header(cell)
        assert cell.font.bold is True

    def test_header_has_center_alignment(self):
        wb = Workbook()
        cell = wb.active.cell(row=1, column=1, value="Header")
        SheetStyler().style_header(cell)
        assert cell.alignment.horizontal == "center"

    def test_header_has_thin_border_all_sides(self):
        wb = Workbook()
        cell = wb.active.cell(row=1, column=1, value="Header")
        SheetStyler().style_header(cell)
        for side in ("top", "bottom", "left", "right"):
            assert getattr(cell.border, side).style == "thin", f"{side} side not thin"

    def test_header_has_gray_fill(self):
        wb = Workbook()
        cell = wb.active.cell(row=1, column=1, value="Header")
        SheetStyler().style_header(cell)
        # PatternFill(start_color="DDDDDD") stores start_color as a Color
        # whose .rgb is "00DDDDDD" (alpha + hex) in openpyxl 3.x.
        assert cell.fill.fill_type == "solid"
        rgb = str(cell.fill.start_color.rgb)
        assert rgb.upper().endswith("DDDDDD")


class TestStyleDataCell:
    """Data cells: thin border + configurable alignment + optional number format."""

    def test_data_cell_default_left_alignment(self):
        wb = Workbook()
        cell = wb.active.cell(row=2, column=1, value="hello")
        SheetStyler().style_data_cell(cell)
        assert cell.alignment.horizontal == "left"

    def test_data_cell_thin_border_all_sides(self):
        wb = Workbook()
        cell = wb.active.cell(row=2, column=1, value="hello")
        SheetStyler().style_data_cell(cell)
        for side in ("top", "bottom", "left", "right"):
            assert getattr(cell.border, side).style == "thin"

    def test_data_cell_right_alignment_override(self):
        wb = Workbook()
        cell = wb.active.cell(row=2, column=1, value=42)
        SheetStyler().style_data_cell(cell, align="right")
        assert cell.alignment.horizontal == "right"

    def test_data_cell_number_format_set(self):
        wb = Workbook()
        cell = wb.active.cell(row=2, column=1, value=1234.5)
        SheetStyler().style_data_cell(cell, number_format="#,##0.00")
        assert cell.number_format == "#,##0.00"


class TestAutoAdjustWidths:
    """Column-width auto-adjust: (max_content_len + 2) * 1.2."""

    def test_auto_adjust_width_uses_longest_value(self):
        wb = Workbook()
        ws = wb.active
        # Column A: longest value is "SKS-2026-00001" (14 chars)
        ws["A1"].value = "Header"
        ws["A2"].value = "SKS-2026-00001"
        SheetStyler().auto_adjust_widths(ws)
        expected = (14 + 2) * 1.2
        assert abs(ws.column_dimensions["A"].width - expected) < 0.001

    def test_auto_adjust_width_handles_none_cells(self):
        """None / non-string cells must not crash (suppress path)."""
        wb = Workbook()
        ws = wb.active
        ws["A1"].value = "abc"
        ws["A2"].value = None
        ws["A3"].value = datetime(2026, 1, 1)
        SheetStyler().auto_adjust_widths(ws)
        # Should not raise; width is set to some positive number
        assert ws.column_dimensions["A"].width is not None
        assert ws.column_dimensions["A"].width > 0

    def test_auto_adjust_width_handles_empty_sheet(self):
        """A worksheet with no cell values must not crash."""
        wb = Workbook()
        ws = wb.active
        SheetStyler().auto_adjust_widths(ws)
        # No error is the assertion
