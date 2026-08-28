import io
import re

from openpyxl import Workbook
from openpyxl.styles import Font

from app.billing.excel_styler import SheetStyler


def format_price(price_str):
    """Format a price string to a float, handling various formats.
    Returns float or 0.0 if parsing fails.
    """
    if not price_str:
        return 0.0

    try:
        # Strip non-numeric characters except decimal point
        cleaned = re.sub(r"[^\d.]", "", str(price_str))
        if not cleaned:
            return 0.0
        return float(cleaned)
    except (ValueError, TypeError):
        return 0.0


def compute_summary(samples):
    """Compute summary statistics from a list of sample records.

    Args:
        samples: List of Sample objects or dicts with keys: sample_type, price

    Returns:
        dict with:
        - by_type: dict {sample_type: {count, total_price}}
        - grand_total: float
        - total_count: int

    """
    by_type = {}
    total_count = 0
    grand_total = 0.0

    for sample in samples:
        if isinstance(sample, dict):
            sample_type = sample.get("sample_type")
            price_str = sample.get("price")
        else:
            sample_type = getattr(sample, "sample_type", None)
            price_str = getattr(sample, "price", None)

        # Normalize empty/None types to 'Other'
        if not sample_type:
            sample_type = "Other"

        price = format_price(price_str)

        by_type.setdefault(sample_type, {"count": 0, "total_price": 0.0})

        by_type[sample_type]["count"] += 1
        by_type[sample_type]["total_price"] += price

        total_count += 1
        grand_total += price

    return {"by_type": by_type, "grand_total": grand_total, "total_count": total_count}


def generate_excel_report(samples, summary, start_date=None, end_date=None):
    """Generate an Excel workbook with two sheets: Samples and Summary.

    Args:
        samples: List of Sample objects (query results)
        summary: Summary dict from compute_summary()
        start_date: Optional start date string for filename
        end_date: Optional end date string for filename

    Returns:
        tuple: (Excel file bytes, filename)

    """
    wb = Workbook()

    # Remove default sheet
    if "Sheet" in wb.sheetnames:
        del wb["Sheet"]

    # === Sheet 1: Samples ===
    ws_samples = wb.create_sheet("Samples")

    styler = SheetStyler()

    # Header row
    headers = [
        "Sample Code",
        "Sample Name",
        "Sample Type",
        "FSO Name",
        "Collection Date",
        "Submission Date",
        "Retailer FSSAI",
        "Retailer Name",
        "Price",
    ]

    for col_num, header in enumerate(headers, 1):
        cell = ws_samples.cell(row=1, column=col_num, value=header)
        styler.style_header(cell)

    # Data rows
    for row_num, sample in enumerate(samples, 2):
        values = [
            getattr(sample, "sample_code", ""),
            getattr(sample, "sample_name", ""),
            getattr(sample, "sample_type", "") or "",
            getattr(sample, "fso_name", ""),
            getattr(sample, "collection_date", ""),
            getattr(sample, "submission_date", "") or "",
            getattr(sample, "retailer_fssai", "") or "",
            getattr(sample, "retailer_name", "") or "",
            getattr(sample, "price", "") or "",
        ]

        for col_num, value in enumerate(values, 1):
            cell = ws_samples.cell(row=row_num, column=col_num, value=str(value) if value else "")
            styler.style_data_cell(cell, align="left")

    styler.auto_adjust_widths(ws_samples)

    # === Sheet 2: Summary ===
    ws_summary = wb.create_sheet("Summary")

    ws_summary.append([])  # Empty row (preserves original row layout)
    ws_summary.append(["Billing Summary"])
    ws_summary.cell(row=2, column=1, value="Billing Summary").font = Font(bold=True, size=16)

    # Filter info
    ws_summary.append([])
    filter_info = []
    if start_date and end_date:
        filter_info = [f"Date Range: {start_date} to {end_date}"]
    elif start_date:
        filter_info = [f"From: {start_date}"]
    elif end_date:
        filter_info = [f"To: {end_date}"]
    else:
        filter_info = ["All Dates"]

    ws_summary.append(filter_info)
    ws_summary.append([])

    # Summary table headers
    summary_headers = ["Sample Type", "Count", "Total Price"]
    for col_num, header in enumerate(summary_headers, 1):
        cell = ws_summary.cell(row=6, column=col_num, value=header)
        styler.style_header(cell)

    # Summary data rows
    by_type = summary.get("by_type", {})
    for row_num, (sample_type, data) in enumerate(by_type.items(), 7):
        count = data.get("count", 0)
        total_price = data.get("total_price", 0.0)

        values = [sample_type, count, total_price]
        for col_num, value in enumerate(values, 1):
            cell = ws_summary.cell(row=row_num, column=col_num, value=value)
            styler.style_data_cell(
                cell,
                align="left" if isinstance(value, str) else "right",
                number_format="#,##0.00" if col_num == 3 and isinstance(value, (int, float)) else None,
            )

    # Grand total row
    grand_total_row = len(by_type) + 7
    ws_summary.append([])
    grand_total = summary.get("grand_total", 0.0)
    total_count = summary.get("total_count", 0)

    ws_summary.cell(row=grand_total_row, column=1, value="GRAND TOTAL")
    ws_summary.cell(row=grand_total_row, column=2, value=total_count)
    ws_summary.cell(row=grand_total_row, column=3, value=grand_total)

    for col_num in range(1, 4):
        cell = ws_summary.cell(row=grand_total_row, column=col_num)
        styler.style_total_row(
            cell,
            is_total_label=(col_num == 1),
            number_format="#,##0.00" if col_num == 3 else None,
        )

    styler.auto_adjust_widths(ws_summary)

    # === Save to BytesIO ===
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)

    # Generate filename
    if start_date and end_date:
        filename = f"billing_summary_{start_date}_to_{end_date}.xlsx"
    elif start_date:
        filename = f"billing_summary_{start_date}_to_end.xlsx"
    elif end_date:
        filename = f"billing_summary_start_to_{end_date}.xlsx"
    else:
        filename = "billing_summary_all.xlsx"

    # Clean filename (replace slashes with underscores)
    filename = filename.replace("/", "_").replace("\\", "_")

    return output, filename
